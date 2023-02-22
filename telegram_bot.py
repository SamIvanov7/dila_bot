import logging
import re

import telebot
from openpyxl import load_workbook

bot = telebot.TeleBot("6038351834:AAHrRDBUteglTmvBGPEQjN1mor8_8Bm1Blk")
file_path = "excel_docs/S.xlsx"

# Set up logging
logging.basicConfig(
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s", level=logging.INFO
)


@bot.message_handler(content_types="text", commands=["send"])
def send(message):
    """Handle the /send command to send the Excel file."""
    global file_path
    bot.send_document(chat_id=message.chat.id, document=open(file_path, "rb"))


@bot.message_handler(content_types=["document"])
def receive(message):
    """Handle a document sent to the bot and save it to file."""
    global file_path
    file_id = message.document.file_id
    bot.download_file(file_id, file_path)


@bot.message_handler(content_types="text", commands=["delete"])
def delete_last_row(message):
    """Handle the /delete command to delete the last row of the Excel file."""
    global file_path
    try:
        # Open the Excel file using openpyxl
        wb = load_workbook(filename=file_path)
    except FileNotFoundError as e:
        print(f"Error opening file {file_path}: {e}")
        return

    try:
        # Select the first worksheet
        ws = wb.active
        # Find the last row in the worksheet
        last_row = ws.max_row
        # Delete the last row
        ws.delete_rows(last_row)
        # Save the updated data back to the Excel file
        wb.save(file_path)
        bot.send_message(chat_id=message.chat.id, text="deleted")
    except PermissionError as e:
        print(f"Error saving file {file_path}: {e}")
        return


@bot.message_handler(content_types="text")
def handle_text_message(message):
    """Handle a text message containing data and update the Excel file."""
    global file_path
    data_string = message.text
    text = data_string

    def update_excel_file(file_path):
        """Update the Excel file with the data in the text message."""
        try:
            # Open the Excel file using openpyxl
            wb = load_workbook(filename=file_path)
        except FileNotFoundError as e:
            print(f"Error opening file {file_path}: {e}")
            return

        try:
            # Select the first worksheet
            ws = wb.active
            # Find the last row in the worksheet
            last_row = ws.max_row
            # Add a new row with increasing numbers
            number = last_row + 1
            # Update the columns based on the dictionary
            new_row = [
                number,
                date,
                name,
                city,
                None,
                phone,
                email,
                address,
                None,
                None,
                None,
                None,
            ]
            # Append the new row to the worksheet
            ws.append(new_row)
            # Save the updated data back to the Excel file
            wb.save(file_path)
        except PermissionError as e:
            print(f"Error saving file {file_path}: {e}")
            return

    try:
        digits = "".join(re.findall(r"\d+", text))
        day = digits[:2]
        month = digits[2:4]
        year = digits[4:8]
        date = f"{day}.{month}.{year}"

        name = re.search(r"Ваше_імя_прізвище_по_батькові_: (.+?)\n", text).group(1)
        name = name if name else None
        city = re.search(
            r"В_якому_місті_Ви_плануєте_відкрити_діагностичне_відділення_МЛ_ДІЛА_: (.*)",
            text,
        )
        city = city.group(1) if city else None
        result = str(re.findall(r"Phone:.\w+", text))
        result = result.split(" ")
        result = result[1]
        result = result[:10]
        phone = result
        phone1 = digits[8:18]
        email = re.search(r"Email: (.*)", text)
        email = email.group(1) if email else None
        address = re.search(
            r"Якщо_у_Вас_є_приміщення_в_якому_ви_бажаєте_розмістити_франчайзингове_відділення_вкажіть_повну_адресу: (.*)",
            text,
        )
        address = address.group(1) if address else None
        update_excel_file(file_path)
        bot.send_message(
            chat_id=message.chat.id, text="Successfully updated the Excel file."
        )
    except Exception as e:
        # Call the error handler function
        print(e, result)


if __name__ == "__main__":
    bot.polling()
