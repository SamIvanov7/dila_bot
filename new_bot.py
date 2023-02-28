import logging
import re

import telebot
from openpyxl import load_workbook

bot = telebot.TeleBot("")

file_path = "excel_docs/S.xlsx"

import pandas as pd

# reg = pd.read_excel("excel_docs/regions.xlsx")
# reg = reg.iloc[:, :2]
# reg = reg.values
# reg_list = reg.tolist()
with open("excel_docs/reg_list.txt", "r") as f:
    reg_list_ru = f.readlines()
list_ru = []
for line in reg_list_ru:
    list_ru.append(line.strip())
list_r = [eval(item) for item in list_ru]
with open("excel_docs/reg_list_ukr.txt", "r") as f:
    reg_list_ukr = f.readlines()
list_ua = []
for line in reg_list_ukr:
    list_ua.append(line.strip())
list_u = [eval(item) for item in list_ua]


def find_region(city, data):
    for row in data:
        if row[0] == city:
            return row[1]
    return None


def find_first_word(text):
    match = re.search(r"\w+", text)
    if match:
        return match.group()
    return None


def digits_to_string(digits):
    string = "".join(digits)
    return string


# Set up logging
logging.basicConfig(
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s", level=logging.INFO
)


@bot.message_handler(content_types="text", commands=["send"])
def send(message):
    """Handle the /send command to send the Excel file."""
    global file_path
    bot.send_document(chat_id=message.chat.id, document=open(file_path, "rb"))


@bot.message_handler(content_types=["document"])
def receive(message):
    """Handle a document sent to the bot and save it to file."""
    global file_path
    file_id = message.document.file_id
    bot.download_file(file_id, file_path)


@bot.message_handler(content_types="text", commands=["delete"])
def delete_last_row(message):
    """Handle the /delete command to delete the last row of the Excel file."""
    global file_path
    try:
        # Open the Excel file using openpyxl
        wb = load_workbook(filename=file_path)
    except FileNotFoundError as e:
        print(f"Error opening file {file_path}: {e}")
        return

    try:
        # Select the first worksheet
        ws = wb.active
        # Find the last row in the worksheet
        last_row = ws.max_row
        # Delete the last row
        ws.delete_rows(last_row)
        # Save the updated data back to the Excel file
        wb.save(file_path)
        bot.send_message(chat_id=message.chat.id, text="deleted")
    except PermissionError as e:
        print(f"Error saving file {file_path}: {e}")
        return


@bot.message_handler(content_types="text")
def handle_text_message(message):
    """Handle a text message containing data and update the Excel file."""
    global file_path
    data_string = message.text
    text = data_string

    def update_excel_file(file_path):
        """Update the Excel file with the data in the text message."""
        try:
            # Open the Excel file using openpyxl
            wb = load_workbook(filename=file_path)
        except FileNotFoundError as e:
            print(f"Error opening file {file_path}: {e}")
            return

        try:
            # Select the first worksheet
            ws = wb.active
            # Find the last row in the worksheet
            last_row = ws.max_row
            # Add a new row with increasing numbers
            number = last_row + 1
            # Update the columns based on the dictionary
            new_row = [
                number,
                date,
                name,
                town,
                region,
                phone,
                email,
                adress,
                None,
                None,
                None,
                None,
            ]
            # Append the new row to the worksheet
            ws.append(new_row)
            # Save the updated data back to the Excel file
            wb.save(file_path)
            return number
        except PermissionError as e:
            print(f"Error saving file {file_path}: {e}")

    def date_search(text):
        date = "".join(re.findall(r"\d+", text))
        if not date:
            date = "28.02.2023"
        return date

    def name_search(text):
        name = re.search(r"Ваше_імя_прізвище_по_батькові_: (.+?)\n", text)
        if not name:
            name = re.search(r"Ім'я - (.+?)\n", text)
        return name

    def city_search(text):
        city = re.search(
            r"В_якому_місті_Ви_плануєте_відкрити_діагностичне_відділення_МЛ_ДІЛА_: (.*)",
            text,
        )
        if not city:
            city = re.search(r"Місто -(.*)", text)
        return city

    def phone_search(text):
        phone = re.search(r"Phone: (.*)", text)
        if not phone:
            phone = re.search(r"Телефон - (.*)", text)
        else:
            None
        return phone

    def email_search(text):
        email = re.search(r"Email: (.*)", text)
        if not email:
            email = re.search(r"E-mail -(.*)", text)
        else:
            None
        return email

    def adress_search(text):
        adress = re.search(
            r"Якщо_у_Вас_є_приміщення_в_якому_ви_бажаєте_розмістити_франчайзингове_відділення_вкажіть_повну_адресу: (.*)",
            text,
        )
        if not adress:
            adress = None
        return adress

    try:
        digits = date_search(text)
        day = digits[:2]
        month = digits[2:4]
        year = digits[4:8]
        date = f"{day}.{month}.{year}"
        name = name_search(text)
        if name is not None:
            name = name.group(1)
        city = city_search(text)
        if city is not None:
            city = city.group(1)
        town = str(find_first_word(city))
        region = find_region(town, list_r)
        if region is None:
            region = find_region(town, list_u)

        phone = phone_search(text)
        if phone is not None:
            phone = phone.group(1)
        phone = digits_to_string(phone)
        email = email_search(text)
        if email is not None:
            email = email.group(1)

        adress = adress_search(text)
        if adress is not None:
            adress = adress.group(1)

        # number = print(f"{update_excel_file(file_path)}")
        bot.send_message(
            chat_id=message.chat.id,
            text=f"В файл добавлена новая строка 8====o \nНомер: {update_excel_file(file_path)}\nДата: {date}\nФИО: {name}\nГород: {town}\nОбласть: {region}\nТелефон: {phone}\nПочта: {email}\nАдрес: {adress}",
        )
    except Exception as e:
        bot.send_message(
            chat_id=message.chat.id,
            text=f"сасай, ошибка! Проверь форму ебаный в рот!   Error: {e}",
        )
        # Call the error handler function
        print(
            e,
        )


if __name__ == "__main__":
    bot.polling()
